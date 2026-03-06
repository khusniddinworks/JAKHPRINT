import logging
import sqlite3
import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from telegram import (
    Update, KeyboardButton, ReplyKeyboardMarkup, ReplyKeyboardRemove, 
    WebAppInfo, InlineKeyboardButton, InlineKeyboardMarkup
)
from telegram.ext import (
    Application, CommandHandler, ConversationHandler, 
    MessageHandler, filters, ContextTypes, CallbackQueryHandler
)
import threading
import http.server
import socketserver
from config import TELEGRAM_TOKEN, ADMIN_ID

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ── Fayl yo'li ───────────────────────────────────────
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "works.xlsx")

# ── States ───────────────────────────────────────────
(SELECT_CATEGORY, SELECT_SUB, ENTER_DETAILS, CONFIRM_STATE, 
 BROADCAST_STATE, CALC_START, CALC_SERVICES, CALC_PLAN, CALC_ADDONS,
 PRICE_EDIT_CAT, PRICE_EDIT_SVC, PRICE_EDIT_VAL, ANSWER_STATE, WAITING_PHONE) = range(14)

# ── Narxlar fayli ─────────────────────────────────
PRICES_FILE = os.path.join(os.path.dirname(__file__), "prices.json")

def load_prices():
    """prices.json dan narxlarni o'qish."""
    with open(PRICES_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_prices(data):
    """prices.json ga narxlarni yozish."""
    with open(PRICES_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

# ── Narxlar va Ma'lumotlar (Dinamik) ───────────────────
PRICES = {}
ADDONS_PRICES = {}
SUB_BUTTONS = {}

def sync_prices():
    """prices.json dan hamma ma'lumotlarni global o'zgaruvchilarga yuklaydi."""
    global PRICES, ADDONS_PRICES, SUB_BUTTONS
    try:
        data = load_prices()
        new_prices = {}
        new_subs = {}
        
        for cat in data["categories"]:
            cat_title = cat["title"]
            services_dict = {}
            subs_list = []
            
            for svc in cat["services"]:
                # Kalkulyator va Bot uchun nom: "Xizmat nomi (Narxi so'm!)"
                # Biz prices.json dagi formatni saqlaymiz, lekin bot menyusi uchun chiroyli ko'rinish beramiz
                display_name = f"{svc['name']} ({svc['price']:,.0f} so'm)".replace(",", " ")
                services_dict[display_name] = svc["price"]
                subs_list.append(display_name)
                
                # Maxsus buyurtma har doim bo'lishi kerak
                if cat["id"] in ["web", "bot"]:
                    if "⚙️ Maxsus buyurtma" not in subs_list:
                        subs_list.append("⚙️ Maxsus buyurtma")
            
            subs_list.append("⬅️ Orqaga")
            new_prices[cat_title] = services_dict
            new_subs[cat_title] = subs_list
            
        PRICES = new_prices
        SUB_BUTTONS = new_subs
        
        # Addonlarni (vizitka/flayer) print kategoriyasidan qidiramiz
        for cat in data["categories"]:
            if cat["id"] == "print":
                for svc in cat["services"]:
                    if "Vizitka" in svc["name"] or "Flayer" in svc["name"]:
                        ADDONS_PRICES[svc["name"]] = svc["price"]
        
        logger.info("✅ Narxlar muvaffaqiyatli sinxronizatsiya qilindi.")
    except Exception as e:
        logger.error(f"❌ Narxlarni sinxronlashda xato: {e}")

# Dastlabki yuklash
sync_prices()

ADMIN_ONLY_BUTTONS = ["📊 Statistika", "📂 Excelni yuklab olish", "📢 Xabar yuborish", "💰 Narxlarni o'zgartirish"]

def get_uzb_time():
    """O'zbekiston vaqtini qaytaradi (UTC+5)."""
    from datetime import timedelta
    return datetime.now() + timedelta(hours=5)

CONFIRM_BUTTONS = ["✅ Tasdiqlash", "✏️ Tahrirlash", "❌ Bekor qilish"]


def make_keyboard(buttons: list, columns: int = 2) -> ReplyKeyboardMarkup:
    """Buttons ro'yxatidan ReplyKeyboardMarkup yasaydi."""
    rows = [buttons[i:i + columns] for i in range(0, len(buttons), columns)]
    return ReplyKeyboardMarkup(
        [[KeyboardButton(b) for b in row] for row in rows],
        resize_keyboard=True,
        is_persistent=True,
    )


CONFIRM_KB = make_keyboard(CONFIRM_BUTTONS, columns=1)


def get_main_keyboard(user_id):
    # Render URL (dinamik narxlar uchun)
    web_app_url = os.environ.get("RENDER_EXTERNAL_URL", "https://khusniddinworks.github.io/JAKHPRINT/")
    
    keyboard = []
    keyboard.append([KeyboardButton("🚀 Xizmatlar va Narxlar", web_app=WebAppInfo(url=web_app_url))])
    keyboard.append([KeyboardButton("📞 Bog'lanish"), KeyboardButton("ℹ️ Biz haqimizda")])
        
    if user_id == ADMIN_ID:
        for i in range(0, len(ADMIN_ONLY_BUTTONS), 2):
            row = [KeyboardButton(b) for b in ADMIN_ONLY_BUTTONS[i:i+2]]
            keyboard.append(row)
            
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, is_persistent=True)


# ── Excel ─────────────────────────────────────────────
HEADERS = ["ID", "Sana", "User ID", "Username", "Kategoriya", "Xizmat", "Tafsilot"]

def init_excel():
    """Faylni tekshiradi va sarlavhalar yo'q bo'lsa qo'shadi."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Zakazlar"
        ws.append(HEADERS)
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        # Agar fayl butkul bo'sh bo'lsa yoki birinchi qator sarlavha bo'lmasa
        if ws.max_row < 1 or (ws.cell(row=1, column=1).value != "ID" and ws.cell(row=1, column=1).value != "#"):
            # Faylni tozalab sarlavha qo'shish
            ws.delete_rows(1, ws.max_row)
            ws.append(HEADERS)
    
    # Ustun kengliklarini yangilash
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 50
    wb.save(EXCEL_FILE)
    logger.info("✅ works.xlsx tayyor holatga keltirildi.")

def save_to_excel(user_id, username, category, service, details):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    last_row = ws.max_row
    next_id = 1
    if last_row > 1:
        val = ws.cell(row=last_row, column=1).value
        try: next_id = int(val) + 1
        except: next_id = last_row
    sana = get_uzb_time().strftime("%Y-%m-%d %H:%M")
    ws.append([next_id, sana, user_id, username, category, service, details])
    wb.save(EXCEL_FILE)
    return next_id

def order_count() -> int:
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    return max(0, ws.max_row - 1)

def init_db():
    conn = sqlite3.connect("orders.db")
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS orders (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER, username TEXT, category TEXT, service TEXT, details TEXT, timestamp DATETIME DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS users (user_id INTEGER PRIMARY KEY, username TEXT, first_name TEXT, last_seen DATETIME DEFAULT CURRENT_TIMESTAMP)""")
    conn.commit()
    conn.close()

def save_user(user_id, username, first_name):
    conn = sqlite3.connect("orders.db")
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO users (user_id, username, first_name, last_seen) VALUES (?, ?, ?, CURRENT_TIMESTAMP)", (user_id, username, first_name))
    conn.commit()
    conn.close()

def get_all_users():
    conn = sqlite3.connect("orders.db")
    c = conn.cursor()
    c.execute("SELECT user_id FROM users")
    users = [row[0] for row in c.fetchall()]
    conn.close()
    return users

def save_to_db(user_id, username, category, service, details):
    conn = sqlite3.connect("orders.db")
    c = conn.cursor()
    c.execute("INSERT INTO orders (user_id, username, category, service, details) VALUES (?,?,?,?,?)", (user_id, username, category, service, details))
    conn.commit()
    conn.close()

# ── Handlers ──────────────────────────────────────────
def run_health_check():
    """Render uchun xavfsiz HTTP server. Faqat ruxsat berilgan fayllarni beradi."""
    port = int(os.environ.get("PORT", 8000))
    web_dir = os.path.dirname(__file__)
    ALLOWED_EXTENSIONS = ('.html', '.css', '.js', '.json', '.png', '.jpg', '.ico', '.svg', '.webp')
    
    class SafeHandler(http.server.SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=web_dir, **kwargs)
        
        def do_GET(self):
            # Query parametrlarni olib tashlash (masalan, ?t=123)
            clean_path = self.path.split('?')[0]
            
            # Faqat ruxsat berilgan fayl turlarini berish
            allowed = (clean_path == '/' or clean_path == '/index.html' or 
                       any(clean_path.endswith(ext) for ext in ALLOWED_EXTENSIONS))
            
            # Debug uchun log
            # print(f"DEBUG: Path {self.path} -> Clean {clean_path} -> Allowed: {allowed}")
            
            if allowed:
                super().do_GET()
            else:
                self.send_response(403)
                self.end_headers()
                self.wfile.write(b'Forbidden')

    with socketserver.TCPServer(("", port), SafeHandler) as httpd:
        logger.info(f"\u2705 Xavfsiz server {port}-portda ishga tushdi.")
        httpd.serve_forever()

async def keep_alive(context: ContextTypes.DEFAULT_TYPE):
    """Botni uxlamasligi uchun har 10 daqiqada getMe so'rovini yuboradi."""
    try:
        await context.bot.get_me()
        logger.info("❇️ Keep-alive: Bot faol.")
    except Exception as e:
        logger.error(f"Keep-alive xatosi: {e}")

async def set_bot_info(context: ContextTypes.DEFAULT_TYPE):
    """Botning tavsifini va ma'lumotlarini o'rnatadi."""
    try:
        # "What can this bot do?" qismi
        await context.bot.set_my_description(
            "🤖 Ushbu bot orqali Veb-sayt yaratish, Telegram botlar yasash va barcha turdagi Print xizmatlariga buyurtma berishingiz mumkin.\n\n"
            "👨‍💻 Admin: @khusniddinkhamidov"
        )
        # Qisqa tavsif
        await context.bot.set_my_short_description(
            "Veb-sayt, Bot va Print xizmatlari uchun rasmiy zakaz boti. Admin: @khusniddinkhamidov"
        )
        logger.info("✅ Bot tavsiflari o'rnatildi.")
    except Exception as e:
        logger.error(f"Bot ma'lumotlarini o'rnatishda xato: {e}")

# ── Admin Buyurtmaga Javob ─────────────────────────
async def order_action_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    data = query.data.split("_") # pattern: "ord_action_userid_orderid"
    action = data[1]
    target_user_id = int(data[2])
    order_id = data[3]
    
    if action == "reply":
        context.user_data["reply_to_user"] = target_user_id
        context.user_data["reply_order_id"] = order_id
        await query.message.reply_text(f"✍️ Buyurtma #{order_id} uchun javobingizni yozing:")
        return ANSWER_STATE
    
    elif action == "done":
        await query.edit_message_caption(caption=query.message.caption + "\n\n✅ *BAJARILDI*", parse_mode="Markdown")
        try:
            await context.bot.send_message(chat_id=target_user_id, text=f"✅ Buyurtmangiz #{order_id} muvaffaqiyatli bajarildi!")
        except: pass

async def answer_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text
    target_user_id = context.user_data.get("reply_to_user")
    order_id = context.user_data.get("reply_order_id")
    
    if not target_user_id: return SELECT_CATEGORY
    
    try:
        msg = f"📩 *Admin javobi (Buyurtma #{order_id}):*\n\n{text}"
        await context.bot.send_message(chat_id=target_user_id, text=msg, parse_mode="Markdown")
        await update.message.reply_text("✅ Javob yuborildi.", reply_markup=get_main_keyboard(ADMIN_ID))
    except Exception as e:
        await update.message.reply_text(f"❌ Xabar yuborishda xato: {e}")
        
    return SELECT_CATEGORY

# ── Narxlarni o'zgartirish Handlers ─────────────────
async def price_edit_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    prices = load_prices()
    kb = [cat["title"] for cat in prices["categories"]] + ["⬅️ Bekor qilish"]
    await update.message.reply_text("Kategoriyani tanlang:", reply_markup=make_keyboard(kb, 1))
    return PRICE_EDIT_CAT

async def price_edit_cat(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text
    if text == "⬅️ Bekor qilish": return await start(update, context)
    
    prices = load_prices()
    category = next((c for c in prices["categories"] if c["title"] == text), None)
    if not category: return PRICE_EDIT_CAT
    
    context.user_data["edit_cat_id"] = category["id"]
    kb = [s["name"] for s in category["services"]] + ["⬅️ Orqaga"]
    await update.message.reply_text(f"{text} uchun xizmatni tanlang:", reply_markup=make_keyboard(kb, 1))
    return PRICE_EDIT_SVC

async def price_edit_svc(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text
    if text == "⬅️ Orqaga": return await price_edit_start(update, context)
    
    cat_id = context.user_data.get("edit_cat_id")
    prices = load_prices()
    cat = next(c for c in prices["categories"] if c["id"] == cat_id)
    svc = next((s for s in cat["services"] if s["name"] == text), None)
    if not svc: return PRICE_EDIT_SVC
    
    context.user_data["edit_svc_name"] = text
    await update.message.reply_text(f"*{text}* uchun yangi narxni kiriting (faqat raqamda):", parse_mode="Markdown", reply_markup=ReplyKeyboardRemove())
    return PRICE_EDIT_VAL

async def price_edit_val(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text
    if not text.isdigit():
        await update.message.reply_text("❌ Iltimos faqat raqam kiriting!")
        return PRICE_EDIT_VAL
    
    new_price = int(text)
    cat_id = context.user_data.get("edit_cat_id")
    svc_name = context.user_data.get("edit_svc_name")
    
    prices = load_prices()
    for cat in prices["categories"]:
        if cat["id"] == cat_id:
            for s in cat["services"]:
                if s["name"] == svc_name:
                    s["price"] = new_price
                    break
    
    save_prices(prices)
    sync_prices() # Global o'zgaruvchilarni yangilash
    await update.message.reply_text(f"✅ *{svc_name}* narxi {new_price:,.0f} so'm qilib belgilandi!".replace(",", " "), parse_mode="Markdown", reply_markup=get_main_keyboard(ADMIN_ID))
    return SELECT_CATEGORY

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    user = update.effective_user
    save_user(user.id, user.username, user.first_name)
    name = user.first_name or "Mehmon"
    
    msg = (
        f"👋 *Assalomu alaykum, {name}!*\n\n"
        "🚀 *Xizmatlar va Narxlar* tugmasini bosib, barcha xizmatlarimizni ko'ring va buyurtma bering!\n\n"
        "👨‍💻 Admin: @khusniddinkhamidov"
    )
    if user.id == ADMIN_ID:
        count = order_count()
        msg += f"\n\n👑 *Admin Panel:* Jami zakazlar: *{count} ta*"

    await update.message.reply_text(
        msg,
        parse_mode="Markdown",
        reply_markup=get_main_keyboard(user.id),
    )
    return SELECT_CATEGORY

async def category_selected(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text
    user_id = update.effective_user.id
    
    # Admin tugmalarini tekshirish
    if user_id == ADMIN_ID:
        if text == "📊 Statistika":
            count = order_count()
            u_count = len(get_all_users())
            await update.message.reply_text(f"📊 *Statistika:*\n\n✅ Jami zakazlar: *{count} ta*\n👥 Jami foydalanuvchilar: *{u_count} ta*", parse_mode="Markdown")
            return SELECT_CATEGORY
        elif text == "📂 Excelni yuklab olish":
            if os.path.exists(EXCEL_FILE):
                with open(EXCEL_FILE, "rb") as f:
                    await update.message.reply_document(document=f, filename="works.xlsx")
            else:
                await update.message.reply_text("❌ Fayl topilmadi.")
            return SELECT_CATEGORY
        elif text == "📢 Xabar yuborish":
            await update.message.reply_text("📢 Barchaga yubormoqchi bo'lgan xabaringizni yozing (yoki /cancel):", reply_markup=ReplyKeyboardRemove())
            return BROADCAST_STATE
        elif text == "💰 Narxlarni o'zgartirish":
            return await price_edit_start(update, context)

    # "Bog'lanish" va "Biz haqimizda" tugmalari
    if text == "📞 Bog'lanish":
        await update.message.reply_text(
            "📞 *Bog'lanish:*\n\n"
            "👨\u200d💻 Admin: @khusniddinkhamidov\n"
            "� Reklama bo'yicha: @TSH\\_Jamshidbek\n"
            "�📱 Telefon: Telegram orqali yozing\n\n"
            "Ish vaqti: 09:00 — 22:00",
            parse_mode="Markdown"
        )
        return SELECT_CATEGORY
    
    if text == "ℹ️ Biz haqimizda":
        await update.message.reply_text(
            "ℹ️ *JAKHPRINT haqida:*\n\n"
            "🌐 Veb-sayt yaratish\n"
            "🤖 Telegram bot ishlab chiqish\n"
            "🖨️ Print xizmatlari (vizitka, flayer, taklifnomalar)\n\n"
            "Biz bilan ishlaganingiz uchun rahmat! �",
            parse_mode="Markdown"
        )
        return SELECT_CATEGORY

    return SELECT_CATEGORY

# ── Hisob-kitob (Kalkulyator) Handlers ──────────────────
async def calculator_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    items = context.user_data.get("calc_items", [])
    cart_text = f"  (🛒 {len(items)} ta)" if items else ""
    
    kb = ["🌐 Veb-sayt xizmatlari", "🤖 Bot xizmatlari", "🖨️ Print xizmatlari", f"✅ Hisoblash{cart_text}", "🗑 Savatni tozalash", "⬅️ Chiqish"]
    await update.message.reply_text(
        "🧮 *Kalkulyator bo'limiga xush kelibsiz!*\n\n"
        "Quyidagi bo'limlardan xizmatlarni tanlang va savatchaga qo'shing:\n"
        f"🛒 Savatda: *{len(items)} ta* xizmat",
        parse_mode="Markdown",
        reply_markup=make_keyboard(kb, columns=1)
    )
    return CALC_SERVICES

async def calculator_step(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text
    
    if text == "⬅️ Chiqish":
        return await start(update, context)
    
    if text == "🗑 Savatni tozalash":
        context.user_data["calc_items"] = []
        await update.message.reply_text("🗑 Savat tozalandi!")
        return await calculator_start(update, context)
        
    if text == "🌐 Veb-sayt xizmatlari":
        cat_key = next((k for k in SUB_BUTTONS if "Veb" in k), "🌐 Veb-saytlar")
        kb = SUB_BUTTONS.get(cat_key, [])
        await update.message.reply_text("Veb-sayt turi:", reply_markup=make_keyboard(kb, columns=1))
        return CALC_SERVICES
        
    if text == "🤖 Bot xizmatlari":
        cat_key = next((k for k in SUB_BUTTONS if "Bot" in k), "🤖 Telegram Botlar")
        kb = SUB_BUTTONS.get(cat_key, [])
        await update.message.reply_text("Bot turi:", reply_markup=make_keyboard(kb, columns=1))
        return CALC_SERVICES

    if text == "🖨️ Print xizmatlari":
        cat_key = next((k for k in SUB_BUTTONS if "Print" in k), "🖨️ Print Xizmatlari")
        kb = SUB_BUTTONS.get(cat_key, [])
        await update.message.reply_text("Print turi:", reply_markup=make_keyboard(kb, columns=1))
        return CALC_SERVICES

    if text == "⬅️ Orqaga":
        return await calculator_start(update, context)

    # Savatga qo'shish (dublikat tekshiruvi bilan)
    all_prices = {}
    for cat in PRICES.values():
        all_prices.update(cat)
        
    if text in all_prices:
        items = context.user_data.setdefault("calc_items", [])
        if text in items:
            await update.message.reply_text(f"⚠️ *{text}* allaqachon savatda bor!", parse_mode="Markdown")
        else:
            items.append(text)
            await update.message.reply_text(f"✅ *{text}* savatga qo'shildi. (🛒 {len(items)} ta)", parse_mode="Markdown")
        return CALC_SERVICES

    if "✅ Hisoblash" in text:
        items = context.user_data.get("calc_items", [])
        if not items:
            await update.message.reply_text("⚠️ Savatchangiz bo'sh! Avval xizmat tanlang.")
            return await calculator_start(update, context)
        
        total = sum(all_prices.get(i, 0) for i in items)
        context.user_data["calc_total"] = total
        
        msg = "🛒 *Siz tanlagan xizmatlar:*\n\n"
        for idx, i in enumerate(items, 1):
            price = all_prices.get(i, 0)
            msg += f"{idx}. {i} — {price:,.0f} so'm\n".replace(",", " ")
        msg += f"\n💰 *Jami:* {total:,.0f} so'm".replace(",", " ")
        msg += "\n\n🎁 *DIQQAT: Barcha xizmatlarga 10% chegirma e'lon qilindi!*"
        
        await update.message.reply_text(
            msg + "\n\nDavom etishni istaysizmi?", 
            parse_mode="Markdown", 
            reply_markup=make_keyboard(["✅ Davom etish", "❌ Bekor qilish"], columns=1)
        )
        return CALC_PLAN # Use existing state but modified flow

async def plan_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text
    if text == "❌ Bekor qilish":
        return await start(update, context)
    
    if text == "✅ Davom etish":
        # Skip tariff selection and go to result (or addons if needed)
        # Check if addons are needed
        await update.message.reply_text(
            "Bunga qo'shimcha mahsulotlar ham qo'shamizmi?\n(Ixtiyoriy, tanlamasangiz 'O'tkazib yuborish'ni bosing)",
            reply_markup=make_keyboard(["💳 Vizitka (100 ta)", "📰 Flayer (100 ta)", "➡️ O'tkazib yuborish"], columns=1)
        )
        return CALC_ADDONS
    return CALC_PLAN

async def addons_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text
    context.user_data.setdefault("calc_addons", {})
    
    if text == "➡️ O'tkazib yuborish":
        return await final_calc_result(update, context)
        
    if text in ADDONS_PRICES:
        item = text
        current = context.user_data["calc_addons"].get(item, 0)
        context.user_data["calc_addons"][item] = current + 100
        await update.message.reply_text(
            f"✅ {item} +100 ta qo'shildi (Jami: {context.user_data['calc_addons'][item]} ta).\n\n"
            "Yana qo'shasizmi yoki davom etamizmi?",
            reply_markup=make_keyboard(["💳 Vizitka (100 ta)", "📰 Flayer (100 ta)", "✅ Davom etish"], columns=1)
        )
        return CALC_ADDONS
    
    if text == "✅ Davom etish":
        return await final_calc_result(update, context)
    return CALC_ADDONS

async def final_calc_result(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    total = context.user_data.get("calc_total", 0)
    plan = context.user_data.get("calc_plan", "")
    items = context.user_data.get("calc_items", [])
    addons = context.user_data.get("calc_addons", {})

    # Final price summary (tariffs removed as requested)
    final_total = total + addons_total
    
    summary = (
        f"🏁 *YAKUNIY BUYURTMA HISOBOTI*\n\n"
        "🔥 *10% CHEGIRMA O'RNATILGAN!* 🎁\n\n"
        f"📦 *Xizmatlar:* {len(items)} ta\n"
        f"💰 *Asosiy xizmatlar:* {total:,.0f} so'm\n"
    ).replace(",", " ")
    
    if addons_text:
        summary += f"\n➕ *Qo'shimchalar:*\n{addons_text}"
        
    summary += f"\n🏆 *TO'LANADIGAN JAMI:* {final_total:,.0f} so'm\n\n".replace(",", " ")
    summary += "Ushbu buyurtmani tasdiqlaysizmi?"
    
    context.user_data["final_summary"] = summary
    context.user_data["final_price"] = final_total
    
    await update.message.reply_text(summary, parse_mode="Markdown", reply_markup=CONFIRM_KB)
    return CONFIRM_STATE

async def broadcast_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Admin yozgan xabarni barcha foydalanuvchilarga yuboradi."""
    msg = update.message
    admin_id = update.effective_user.id
    
    if admin_id != ADMIN_ID:
        return SELECT_CATEGORY

    users = get_all_users()
    count = 0
    failed = 0
    await update.message.reply_text(f"🚀 {len(users)} ta foydalanuvchiga xabar yuborish boshlandi...")

    import asyncio
    for uid in users:
        try:
            await context.bot.copy_message(chat_id=uid, from_chat_id=admin_id, message_id=msg.message_id)
            count += 1
        except Exception:
            failed += 1
            continue
        await asyncio.sleep(0.05)  # Telegram rate limit himoyasi
    
    await update.message.reply_text(
        f"✅ Xabar {count} ta foydalanuvchiga yetkazildi.\n"
        f"{'❌ ' + str(failed) + ' ta yuborilmadi.' if failed else ''}",
        reply_markup=get_main_keyboard(admin_id)
    )
    return SELECT_CATEGORY

async def sub_selected(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text
    if text == "⬅️ Orqaga":
        return await start(update, context)

    # Global SUB_BUTTONS dagi kategoriyani topish
    category = context.user_data.get("category", "")
    matching_cat = next((k for k in SUB_BUTTONS if k == category or category in k), None)
    
    if matching_cat and text in SUB_BUTTONS.get(matching_cat, []):
        context.user_data["service"] = text
        context.user_data["category"] = matching_cat # Nomni to'g'rilab saqlash
        
        msg = f"*{text}* tanlandi ✅\n\n✍️ Buyurtmangiz haqida batafsil yozing:"
        if "Maxsus buyurtma" in text or "A4 formatdagi boshqa print" in text:
            msg += "\n\n💡 _Ushbu yo'nalishda buyurtmangizni *Ovozli xabar (Audio)* orqali ham yuborishingiz mumkin!_"
            
        await update.message.reply_text(
            msg,
            parse_mode="Markdown",
            reply_markup=get_main_keyboard(update.effective_user.id),
        )
        return ENTER_DETAILS
    return SELECT_SUB

async def voice_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Ovozli xabarlarni qabul qilish va adminga yuborish."""
    service = context.user_data.get("service", "")
    valid_voice_services = ["Maxsus buyurtma", "A4 formatdagi boshqa print"]
    
    if not any(s in service for s in valid_voice_services):
        await update.message.reply_text(
            "❌ Bu bo'limda faqat matnli xabar qabul qilinadi.",
            reply_markup=get_main_keyboard(update.effective_user.id)
        )
        return SELECT_CATEGORY

    user = update.effective_user
    username = user.username or user.first_name
    category = context.user_data.get("category", "")
    sana = get_uzb_time().strftime('%Y-%m-%d %H:%M')

    # Excel va DB ga saqlash
    order_id = save_to_excel(user.id, username, category, service, "Ovozli xabar yuborildi")
    save_to_db(user.id, username, category, service, "Ovozli xabar yuborildi")

    # Adminga audio va info yuborish
    admin_msg = (
        f"🎙 *YANGI OVOZLI BUYURTMA #{order_id}*\n\n"
        f"👤 *Mijoz:* {user.mention_markdown(name=username)}\n"
        f"🆔 *ID:* `{user.id}`\n"
        f"🗂 *Kategoriya:* {category}\n"
        f"📌 *Xizmat:* {service}\n"
        f"📅 *Sana:* {sana}"
    )
    
    inline_kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("💬 Javob yozish", callback_data=f"ord_reply_{user.id}_{order_id}")],
        [InlineKeyboardButton("✅ Bajarildi", callback_data=f"ord_done_{user.id}_{order_id}")]
    ])
    
    try:
        await context.bot.send_message(chat_id=ADMIN_ID, text=admin_msg, parse_mode="Markdown", reply_markup=inline_kb)
        await context.bot.send_voice(chat_id=ADMIN_ID, voice=update.message.voice.file_id)
    except Exception as e:
        logger.error(f"Admin xabar yuborishda xato: {e}")

    await update.message.reply_text(
        "✅ *Ovozli buyurtmangiz qabul qilindi!*\nTez orada bog'lanamiz. 🙏",
        parse_mode="Markdown",
        reply_markup=get_main_keyboard(user.id),
    )
    return SELECT_CATEGORY

async def enter_details(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["details"] = update.message.text
    category = context.user_data.get("category", "")
    service  = context.user_data.get("service", "")

    await update.message.reply_text(
        f"📋 *Buyurtma ma'lumotlari:*\n\n"
        f"🗂 *Kategoriya:* {category}\n"
        f"📌 *Xizmat:* {service}\n"
        f"📝 *Tafsilot:* {update.message.text}\n\n"
        "Tasdiqlaysizmi?",
        parse_mode="Markdown",
        reply_markup=CONFIRM_KB,
    )
    return CONFIRM_STATE

async def confirm_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text
    user = update.effective_user
    
    if text == "✅ Tasdiqlash":
        summary = context.user_data.get("final_summary")
        if summary:
            # Kalkulyator buyurtmasi
            category = "HISOB-KITOB"
            service = context.user_data.get("calc_plan", "Kalkulyator")
            details = summary.replace("*", "").replace("_", "")
        else:
            category = context.user_data.get("category")
            service  = context.user_data.get("service")
            details  = context.user_data.get("details")
        
        if not category or not service:
            return SELECT_CATEGORY

        username = user.username or user.first_name
        context.user_data.clear()

        order_id = save_to_excel(user.id, username, category, service, details)
        save_to_db(user.id, username, category, service, details)

        admin_warning = (
            f"🔔 *YANGI TASDIQLANGAN BUYURTMA #{order_id}*\n\n"
            f"👤 *Mijoz:* {user.mention_markdown(name=username)}\n"
            f"🆔 *ID:* `{user.id}`\n"
            f"🗂 *Kategoriya:* {category}\n"
            f"📌 *Xizmat:* {service}\n"
            f"📝 *Tafsilot:* \n{details}\n"
            f"📅 *Sana:* {get_uzb_time().strftime('%Y-%m-%d %H:%M')}"
        )
        inline_kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("💬 Javob yozish", callback_data=f"ord_reply_{user.id}_{order_id}")],
            [InlineKeyboardButton("✅ Bajarildi", callback_data=f"ord_done_{user.id}_{order_id}")]
        ])
        
        try:
            await context.bot.send_message(chat_id=ADMIN_ID, text=admin_warning, parse_mode="Markdown", reply_markup=inline_kb)
        except Exception as e:
            logger.error(f"Admin xabar yuborishda xato: {e}")

        await update.message.reply_text(
            f"✅ *Buyurtmangiz qabul qilindi!*\nTez orada bog'lanamiz. 🙏",
            parse_mode="Markdown",
            reply_markup=get_main_keyboard(user.id),
        )
        return SELECT_CATEGORY
    elif text == "✏️ Tahrirlash":
        await update.message.reply_text("✏️ Yangi tafsilotlarni yozing:", reply_markup=ReplyKeyboardRemove())
        return ENTER_DETAILS
    else:
        await update.message.reply_text("❌ Bekor qilindi.", reply_markup=get_main_keyboard(user.id))
        return SELECT_CATEGORY

async def web_app_data_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Mini Appdan kelgan ma'lumotlarni qayta ishlash."""
    try:
        data = json.loads(update.effective_message.web_app_data.data)
        services = data.get("services", [])
        total = data.get("total", 0)
        user = update.effective_user
        
        if not services:
            await update.message.reply_text("⚠️ Xizmat tanlanmadi.", reply_markup=get_main_keyboard(user.id))
            return SELECT_CATEGORY
        
        details = "🛒 Mini App orqali tanlangan xizmatlar:\n"
        for s in services:
            details += f"• {s}\n"
        details += f"\n💰 Jami: {total:,.0f} so'm".replace(",", " ")

        # Ma'lumotlarni vaqtincha saqlash
        context.user_data["pending_order"] = {
            "category": "MINI APP",
            "service": "Ko'p tarmoqli",
            "details": details
        }

        # Telefon raqamini so'rash
        kb = [[KeyboardButton("📱 Telefon raqamni yuborish", request_contact=True)]]
        await update.message.reply_text(
            "📞 Buyurtmani yakunlash uchun telefon raqamingizni yuboring:",
            reply_markup=ReplyKeyboardMarkup(kb, resize_keyboard=True, one_time_keyboard=True)
        )
        return WAITING_PHONE
        
    except Exception as e:
        logger.error(f"Mini App xatosi: {e}")
        await update.message.reply_text(
            "❌ Xatolik yuz berdi. Iltimos, qaytadan urinib ko'ring.",
            reply_markup=get_main_keyboard(update.effective_user.id)
        )
        return SELECT_CATEGORY

async def contact_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Foydalanuvchi yuborgan telefon raqamini qabul qilish va buyurtmani yakunlash."""
    contact = update.message.contact
    user = update.effective_user
    username = user.username or user.first_name
    phone = contact.phone_number
    
    order_data = context.user_data.get("pending_order")
    if not order_data:
        await update.message.reply_text("❌ Buyurtma ma'lumotlari topilmadi.", reply_markup=get_main_keyboard(user.id))
        return SELECT_CATEGORY

    category = order_data["category"]
    service = order_data["service"]
    details = order_data["details"]
    
    # Telefon raqamini tafsilotlarga qo'shish
    full_details = f"{details}\n\n📞 Tel: {phone}"

    order_id = save_to_excel(user.id, username, category, service, full_details)
    save_to_db(user.id, username, category, service, full_details)
    save_user(user.id, user.username, user.first_name)

    admin_msg = (
        f"🌟 *YANGI TASDIQLANGAN BUYURTMA #{order_id}*\n\n"
        f"👤 *Mijoz:* {user.mention_markdown(name=username)}\n"
        f"🆔 *ID:* `{user.id}`\n"
        f"📞 *Tel:* {phone}\n"
        f"📝 *Tafsilot:* \n{details}\n"
        f"📅 *Sana:* {get_uzb_time().strftime('%Y-%m-%d %H:%M')}"
    )
    
    inline_kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("💬 Javob yozish", callback_data=f"ord_reply_{user.id}_{order_id}")],
        [InlineKeyboardButton("✅ Bajarildi", callback_data=f"ord_done_{user.id}_{order_id}")]
    ])
    
    try:
        await context.bot.send_message(chat_id=ADMIN_ID, text=admin_msg, parse_mode="Markdown", reply_markup=inline_kb)
    except Exception as e:
        logger.error(f"Admin xabar yuborishda xato: {e}")

    await update.message.reply_text(
        f"✅ *Buyurtmangiz qabul qilindi!*\n\n{details}\n\nTez orada bog'lanamiz. 🙏",
        parse_mode="Markdown",
        reply_markup=get_main_keyboard(user.id)
    )
    
    # Ma'lumotlarni tozalash
    context.user_data.pop("pending_order", None)
    return SELECT_CATEGORY

def main():
    init_db()
    init_excel()
    
    # Health Check serverni alohida thread'da ishga tushirish
    threading.Thread(target=run_health_check, daemon=True).start()
    
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    
    # Bot ma'lumotlarini o'rnatish
    app.job_queue.run_once(set_bot_info, when=0)
    # Anti-sleep: har 10 daqiqada bot o'zini "uyg'otadi"
    app.job_queue.run_repeating(keep_alive, interval=600, first=10)

    conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            CallbackQueryHandler(order_action_callback, pattern="^ord_"),
            MessageHandler(filters.StatusUpdate.WEB_APP_DATA, web_app_data_handler)
        ],
        states={
            SELECT_CATEGORY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, category_selected),
                CallbackQueryHandler(order_action_callback, pattern="^ord_"),
                MessageHandler(filters.StatusUpdate.WEB_APP_DATA, web_app_data_handler)
            ],
            SELECT_SUB: [MessageHandler(filters.TEXT & ~filters.COMMAND, sub_selected)],
            ENTER_DETAILS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_details),
                MessageHandler(filters.VOICE, voice_handler),
            ],
            CONFIRM_STATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, confirm_handler)],
            BROADCAST_STATE: [MessageHandler(filters.ALL & ~filters.COMMAND, broadcast_handler)],
            CALC_SERVICES: [MessageHandler(filters.TEXT & ~filters.COMMAND, calculator_step)],
            CALC_PLAN: [MessageHandler(filters.TEXT & ~filters.COMMAND, plan_handler)],
            CALC_ADDONS: [MessageHandler(filters.TEXT & ~filters.COMMAND, addons_handler)],
            PRICE_EDIT_CAT: [MessageHandler(filters.TEXT & ~filters.COMMAND, price_edit_cat)],
            PRICE_EDIT_SVC: [MessageHandler(filters.TEXT & ~filters.COMMAND, price_edit_svc)],
            PRICE_EDIT_VAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, price_edit_val)],
            ANSWER_STATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, answer_handler)],
            WAITING_PHONE: [MessageHandler(filters.CONTACT, contact_handler)],
        },
        fallbacks=[CommandHandler("cancel", lambda u, c: start(u, c))],
        allow_reentry=True,
    )
    app.add_handler(conv)
    # Mini App ma'lumotlari uchun handler (convdan tashqarida ham qolsin har ehtimolga qarshi)
    app.add_handler(MessageHandler(filters.StatusUpdate.WEB_APP_DATA, web_app_data_handler))
    app.run_polling()

if __name__ == "__main__":
    import asyncio
    try:
        # Python 3.10+ da event loop bilan bog'liq muammolarni oldini olish
        try:
            loop = asyncio.get_event_loop()
        except RuntimeError:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
        main()
    except (KeyboardInterrupt, SystemExit):
        logger.info("Bot to'xtatildi.")
    except Exception as e:
        logger.error(f"Kutilmagan xato: {e}")
